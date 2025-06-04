from flask import Blueprint, request, jsonify, send_file  # type: ignore
from extensions import db , socketio
from models.production import ProductionOrder, Batch, BatchMaterialDispensing
from models.user import User  # ✅ Needed for username and validation
from flask_jwt_extended import jwt_required, get_jwt_identity  # type: ignore
from routes.user_routes import role_required
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from barcode import Code128
from barcode.writer import ImageWriter
from PIL import Image as PILImage
import os, io, tempfile
import traceback
from models.recipe import RecipeMaterial

production_bp = Blueprint("production", __name__)


@production_bp.route("/production_orders/export/barcodes", methods=["GET"])
def export_production_orders_excel_with_barcodes():
    try:
        orders = ProductionOrder.query.all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Production Order Barcodes"
        ws.append(["Order Number", "Barcode ID", "Scannable Barcode"])

        row_number = 2

        for order in orders:
            if order.barcode_id:
                barcode_id = order.barcode_id
                try:
                    temp_dir = tempfile.gettempdir()
                    filename = f"{barcode_id}"
                    filepath = os.path.join(temp_dir, f"{filename}.png")

                    code128 = Code128(barcode_id, writer=ImageWriter())
                    code128.save(filepath)

                    image = PILImage.open(filepath)
                    image = image.resize((200, 60))
                    image.save(filepath)

                    ws.cell(row=row_number, column=1, value=order.order_number)
                    ws.cell(row=row_number, column=2, value=barcode_id)

                    img = ExcelImage(filepath)
                    img.width = 150
                    img.height = 50
                    ws.add_image(img, f"C{row_number}")

                    os.remove(filepath)
                    row_number += 1

                except Exception as e:
                    print(f"Failed to generate barcode for {barcode_id}: {e}")
                    ws.cell(row=row_number, column=1, value=order.order_number)
                    ws.cell(row=row_number, column=2, value=barcode_id)
                    row_number += 1

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        return send_file(
            stream,
            download_name="production_orders_with_barcodes.xlsx",
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        import traceback

        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
@production_bp.route("/production_orders", methods=["POST"])
@jwt_required(locations=["headers"])
@role_required(["admin", "operator"])
def create_production_order():
    from models.recipe import RecipeMaterial
    from extensions import socketio
    import traceback

    data = request.get_json()
    print("Received data:", data)

    required_fields = ["order_number", "recipe_id", "batch_size", "scheduled_date"]
    missing = [field for field in required_fields if field not in data]
    if missing:
        return jsonify({"error": f"Missing fields: {', '.join(missing)}"}), 400

    current_user_id = get_jwt_identity()

    try:
        # Step 1: Fetch materials and calculate dosing
        recipe_materials = RecipeMaterial.query.filter_by(recipe_id=data["recipe_id"]).all()
        if not recipe_materials:
            return jsonify({"error": "No materials found for the selected recipe"}), 400

        batch_size = int(data["batch_size"])

        # Total set point per batch
        total_per_batch = sum(float(m.set_point or 0) for m in recipe_materials)

        # Total set point for all batches
        total_for_all_batches = total_per_batch * batch_size

        # Dosing value = average per batch
        dosing_value = round(total_for_all_batches / batch_size, 2)

        # Step 2: Create Production Order with dosing value
        new_order = ProductionOrder(
            order_number=data["order_number"],
            recipe_id=data["recipe_id"],
            batch_size=batch_size,
            scheduled_date=data["scheduled_date"],
            status="planned",
            created_by=current_user_id,
            notes=data.get("notes"),
            barcode_id=data.get("barcode_id"),
            dosing=dosing_value  # ✅ New field
        )
        db.session.add(new_order)

        # Step 3: Update RecipeMaterial set_point *= batch_size
        for material in recipe_materials:
            original = material.set_point or 0
            material.set_point = float(original) * batch_size

        # Step 4: Commit all DB changes
        db.session.commit()

        # Step 5: Emit SocketIO event to notify frontend to soft-refresh
        socketio.emit("order_created", {
            "recipe_id": new_order.recipe_id,
            "order_number": new_order.order_number
        })

        return jsonify({"message": "Production order created, materials updated, dosing calculated!"}), 201

    except Exception as e:
        db.session.rollback()
        print("Exception:", e)
        traceback.print_exc()
        if "Duplicate entry" in str(e):
            return jsonify({"error": "Duplicate order/barcode"}), 400
        return jsonify({"error": "Failed to create order", "details": str(e)}), 500

@production_bp.route("/production_orders/<int:order_id>", methods=["PUT"])
@jwt_required(locations=["headers"])
@role_required(["admin", "operator"])
def update_production_order(order_id):
    from extensions import socketio
    import traceback
    from models.recipe import RecipeMaterial
    from models.recipe import Recipe
    from models.material import Material

    data = request.get_json()
    if not data:
        return jsonify({"error": "No data received"}), 400

    order = ProductionOrder.query.get(order_id)
    if not order:
        return jsonify({"error": "Production order not found"}), 404

    try:
        old_status = order.status

        # Update fields
        if "order_number" in data:
            order.order_number = data["order_number"]
        if "recipe_id" in data:
            order.recipe_id = data["recipe_id"]
        if "batch_size" in data:
            try:
                order.batch_size = float(data["batch_size"])
            except ValueError:
                return jsonify({"error": "Invalid batch_size format"}), 400
        if "scheduled_date" in data:
            order.scheduled_date = data["scheduled_date"]
        if "status" in data:
            new_status = data["status"]
            order.status = new_status

            # ✅ Ensure only one verified order at a time
            if new_status == "verified":
                other_verified_orders = (
                    ProductionOrder.query
                    .filter(ProductionOrder.order_id != order_id, ProductionOrder.status == "verified")
                    .all()
                )
                for o in other_verified_orders:
                    o.status = "pending"
                db.session.commit()

        if "created_by" in data:
            order.created_by = data["created_by"]
        if "notes" in data:
            order.notes = data["notes"]

        db.session.commit()

        # Emit order update
        socketio.emit("order_updated", {
            "recipe_id": order.recipe_id,
            "order_number": order.order_number
        })

        # Emit active materials if verified
        if old_status != "verified" and order.status == "verified":
            pending = (
                db.session.query(RecipeMaterial)
                .filter(
                    RecipeMaterial.recipe_id == order.recipe_id,
                    RecipeMaterial.status == "pending"
                )
                .all()
            )

            materials_data = []
            for mat in pending:
                materials_data.append({
                    "material_id": mat.material_id,
                    "material_name": mat.material.title if mat.material else "—",
                    "barcode": mat.material.barcode_id if mat.material else "",
                    "set_point": mat.set_point,
                    "actual": mat.actual,
                    "margin": mat.margin,
                    "status": mat.status
                })

            recipe = Recipe.query.get(order.recipe_id)

            socketio.emit("active_recipe_materials", {
                "recipe_id": order.recipe_id,
                "recipe_name": recipe.name if recipe else "Unknown",
                "materials": materials_data
            }, namespace='/')

        return jsonify({"message": "Production order updated successfully!"}), 200

    except Exception as e:
        db.session.rollback()
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@production_bp.route("/production_orders/<int:order_id>", methods=["DELETE"])
def delete_production_order(order_id):
    from extensions import socketio
    import traceback

    try:
        # ✅ Step 1: Fetch the production order
        order = ProductionOrder.query.get(order_id)
        if not order:
            return jsonify({"error": "Production order not found"}), 404

        # ✅ Step 2: Prevent deletion if a batch exists
        existing_batch = Batch.query.filter_by(order_id=order_id).first()
        if existing_batch:
            return jsonify({
                "error": "Cannot delete the order because batch data exists. Please delete the batch first."
            }), 400

        # ✅ Step 3: Delete the production order
        db.session.delete(order)

        # ✅ Step 4: Notify frontend via SocketIO
        socketio.emit("order_deleted", {
            "order_id": order_id,
            "recipe_id": order.recipe_id,
            "order_number": order.order_number
        })

        # ✅ Step 5: Commit transaction
        db.session.commit()

        return jsonify({
            "message": f"Production order {order_id} deleted successfully."
        }), 200

    except Exception as e:
        db.session.rollback()
        traceback.print_exc()
        return jsonify({"error": "Internal server error", "details": str(e)}), 500



@production_bp.route("/production_orders", methods=["GET"])
@jwt_required()
def get_production_orders():
    try:
        orders = ProductionOrder.query.all()
        result = []

        for order in orders:
            user = User.query.get(order.created_by)
            result.append({
                "order_id": order.order_id,
                "order_number": order.order_number,
                "recipe_id": order.recipe_id,
                "batch_size": str(order.batch_size),
                "scheduled_date": order.scheduled_date.strftime("%Y-%m-%d") if order.scheduled_date else None,
                "status": order.status,
                "created_by": order.created_by,
                "created_by_username": user.username if user else "—",
                "barcode_id": order.barcode_id,
                "dosing": order.dosing  # ✅ Include dosing
            })

        return jsonify(result), 200

    except Exception as e:
        import traceback
        print("❌ Error fetching production orders:", e)
        traceback.print_exc()
        return jsonify({"error": "Failed to fetch production orders"}), 500


@production_bp.route("/production_orders/<int:order_id>", methods=["GET"])
def get_production_order(order_id):
    order = ProductionOrder.query.get(order_id)
    if not order:
        return jsonify({"error": "Production order not found"}), 404

    result = {
        "order_id": order.order_id,
        "order_number": order.order_number,
        "recipe_id": order.recipe_id,
        "batch_size": str(order.batch_size),
        "scheduled_date": order.scheduled_date.strftime("%Y-%m-%d"),
        "status": order.status,
        "created_by": order.created_by,
        "created_by_username": order.creator.username if order.creator else None,
    }
    return jsonify(result), 200


@production_bp.route("/production-orders/<int:order_id>/reject", methods=["PUT"])
@jwt_required()
@role_required(["admin"])
def reject_production_order(order_id):
    order = ProductionOrder.query.get(order_id)
    if not order:
        return jsonify({"error": "Production order not found"}), 404

    order.status = "rejected"
    db.session.commit()
    return jsonify({"message": "Production order rejected successfully"}), 200


### BATCH ROUTES ###
@production_bp.route("/batches", methods=["POST"])
def create_batch():
    data = request.get_json()
    if not all(key in data for key in ["batch_number", "order_id", "operator_id"]):
        return (
            jsonify(
                {
                    "error": "Missing required fields (batch_number, order_id, operator_id)"
                }
            ),
            400,
        )

    order = ProductionOrder.query.get(data["order_id"])
    if not order:
        return (
            jsonify({"error": f"Order with ID {data['order_id']} does not exist"}),
            400,
        )

    operator = User.query.get(data["operator_id"])
    if not operator:
        return (
            jsonify(
                {"error": f"Operator with ID {data['operator_id']} does not exist"}
            ),
            400,
        )

    if Batch.query.filter_by(batch_number=data["batch_number"]).first():
        return (
            jsonify({"error": f"Batch number {data['batch_number']} already exists"}),
            400,
        )

    try:
        new_batch = Batch(
            batch_number=data["batch_number"],
            order_id=data["order_id"],
            operator_id=data["operator_id"],
            status=data.get("status", "pending"),
            notes=data.get("notes"),
        )
        db.session.add(new_batch)
        db.session.commit()
        return (
            jsonify(
                {
                    "message": "Batch created successfully!",
                    "batch_id": new_batch.batch_id,
                }
            ),
            201,
        )
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@production_bp.route("/batches", methods=["GET"])
def get_batches():
    batches = Batch.query.all()
    result = [
        {
            "batch_id": batch.batch_id,
            "batch_number": batch.batch_number,
            "order_id": batch.order_id,
            "status": batch.status,
            "operator_id": batch.operator_id,
            "notes": batch.notes,
            "created_at": batch.created_at,
        }
        for batch in batches
    ]
    return jsonify(result)


@production_bp.route("/batches/<int:batch_id>", methods=["PUT"])
def update_batch(batch_id):
    batch = Batch.query.get_or_404(batch_id)
    data = request.get_json()

    batch.batch_number = data.get("batch_number", batch.batch_number)
    batch.order_id = data.get("order_id", batch.order_id)
    batch.status = data.get("status", batch.status)
    batch.operator_id = data.get("operator_id", batch.operator_id)
    batch.notes = data.get("notes", batch.notes)

    db.session.commit()
    return jsonify({"message": "Batch updated successfully!"})


@production_bp.route("/batches/<int:batch_id>", methods=["DELETE"])
def delete_batch(batch_id):
    batch = Batch.query.get_or_404(batch_id)
    db.session.delete(batch)
    db.session.commit()
    return jsonify({"message": "Batch deleted successfully!"})


### BATCH MATERIAL DISPENSING ROUTES ###
@production_bp.route("/batch_dispensing", methods=["POST"])
def create_batch_dispensing():
    data = request.get_json()
    new_dispensing = BatchMaterialDispensing(
        batch_id=data["batch_id"],
        material_id=data["material_id"],
        planned_quantity=data["planned_quantity"],
        actual_quantity=data.get("actual_quantity"),
        dispensed_by=data["dispensed_by"],
        status=data.get("status", "pending"),
    )
    db.session.add(new_dispensing)
    db.session.commit()
    return jsonify({"message": "Material dispensing record created successfully!"}), 201


@production_bp.route("/batch_dispensing", methods=["GET"])
def get_batch_dispensing():
    dispensing_records = BatchMaterialDispensing.query.all()
    result = [
        {
            "dispensing_id": record.dispensing_id,
            "batch_id": record.batch_id,
            "material_id": record.material_id,
            "planned_quantity": str(record.planned_quantity),
            "actual_quantity": str(record.actual_quantity)
            if record.actual_quantity
            else None,
            "dispensed_by": record.dispensed_by,
            "status": record.status,
        }
        for record in dispensing_records
    ]
    return jsonify(result)
