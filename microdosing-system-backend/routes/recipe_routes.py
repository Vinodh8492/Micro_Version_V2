from flask import Blueprint, request, jsonify, send_file
from extensions import db
from models.recipe import Recipe, RecipeMaterial, DosedRecipeMaterial
from models.production import ProductionOrder
from models.user import User
from sqlalchemy.exc import IntegrityError
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from barcode import Code128
from barcode.writer import ImageWriter
from PIL import Image as PILImage
import io, os, tempfile
from werkzeug.exceptions import BadRequest
import logging
import time
from sqlalchemy.orm import joinedload
from extensions import socketio
from flask_socketio import emit
from models.material import Material



recipe_bp = Blueprint("recipe", __name__)
logging.basicConfig(level=logging.DEBUG)

@recipe_bp.route("/recipes/export/barcodes", methods=["GET"])
def export_recipes_excel_with_barcodes():
    start_time = time.time()
    try:
        recipes = Recipe.query.with_entities(Recipe.name, Recipe.code, Recipe.barcode_id).all()
        wb = Workbook()
        ws = wb.active
        ws.title = "Recipes with Barcodes"
        ws.append(["Name", "Code", "Barcode ID", "Scannable Barcode"])

        row_number = 2
        for recipe in recipes:
            name, code, barcode_id = recipe
            if barcode_id:
                try:
                    temp_dir = tempfile.gettempdir()
                    filename = f"{barcode_id}"
                    filepath = os.path.join(temp_dir, f"{filename}.png")
                    code128 = Code128(barcode_id, writer=ImageWriter())
                    code128.save(filepath)
                    image = PILImage.open(filepath)
                    image = image.resize((200, 60))
                    image.save(filepath)

                    ws.cell(row=row_number, column=1, value=name)
                    ws.cell(row=row_number, column=2, value=code)
                    ws.cell(row=row_number, column=3, value=barcode_id)

                    img = ExcelImage(filepath)
                    img.width = 150
                    img.height = 50
                    ws.add_image(img, f"D{row_number}")

                    os.remove(filepath)
                    row_number += 1

                except Exception as e:
                    logging.error(f"Failed to generate barcode for {barcode_id}: {e}")
                    ws.cell(row=row_number, column=1, value=name)
                    ws.cell(row=row_number, column=2, value=code)
                    ws.cell(row=row_number, column=3, value=barcode_id)
                    row_number += 1

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        end_time = time.time()
        execution_time_ms = round((end_time - start_time) * 1000, 2)

        response = send_file(
            stream,
            download_name="recipes_with_barcodes.xlsx",
            as_attachment=True,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response.headers['X-Execution-Time-MS'] = str(execution_time_ms)
        return response

    except Exception as e:
        logging.error(f"Error exporting barcodes: {str(e)}")
        return jsonify({"error": str(e)}), 500

@recipe_bp.route("/recipes", methods=["GET"])
def get_recipes():
    start_time = time.time()
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 20, type=int)
    pagination = Recipe.query.paginate(page=page, per_page=per_page, error_out=False)
    recipes = pagination.items

    result = [
        {
            "recipe_id": r.recipe_id,
            "name": r.name,
            "code": r.code,
            "description": r.description,
            "version": r.version,
            "status": r.status,
            "created_by": r.created_by,
            "created_at": r.created_at,
            "no_of_materials": r.no_of_materials,
        } for r in recipes
    ]
    end_time = time.time()
    execution_time_ms = round((end_time - start_time) * 1000, 2)

    return jsonify({
        "recipes": result,
        "total": pagination.total,
        "page": pagination.page,
        "pages": pagination.pages,
        "per_page": pagination.per_page,
        "execution_time_ms": execution_time_ms
    })

@recipe_bp.route("/recipes", methods=["POST"])
def create_recipe():
    start_time = time.time()
    data = request.get_json()
    required_fields = ["name", "code", "version", "created_by"]
    for field in required_fields:
        if not data.get(field):
            return jsonify({"error": f"'{field}' is required."}), 400

    status = data.get("status", "Unreleased")
    valid_statuses = ["Released", "Unreleased"]
    if status not in valid_statuses:
        return jsonify({"error": f"Invalid status value: {status}"}), 400

    user = db.session.get(User, data["created_by"])
    if not user:
        return jsonify({"error": "User not found."}), 400

    new_recipe = Recipe(
        name=data["name"],
        code=data["code"],
        description=data.get("description"),
        version=data["version"],
        status=status,
        created_by=data["created_by"],
        barcode_id=data.get("barcode_id"),
        no_of_materials=data.get("no_of_materials")
    )

    db.session.add(new_recipe)
    try:
        db.session.commit()
    except IntegrityError as e:
        db.session.rollback()
        if "Duplicate entry" in str(e.orig):
            return jsonify({"error": "Duplicate entry: code or barcode_id already exists."}), 400
        return jsonify({"error": "Database error occurred."}), 500
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error creating recipe: {str(e)}")
        return jsonify({"error": str(e)}), 500

    end_time = time.time()
    execution_time_ms = round((end_time - start_time) * 1000, 2)
    return jsonify({"message": "Recipe created successfully!", "execution_time_ms": execution_time_ms}), 201

@recipe_bp.route("/recipes/<int:recipe_id>", methods=["GET"])
def get_recipe(recipe_id):
    start_time = time.time()
    recipe = Recipe.query.get(recipe_id)
    if not recipe:
        return jsonify({"error": "Recipe not found"}), 404

    result = {
        "recipe_id": recipe.recipe_id,
        "name": recipe.name,
        "code": recipe.code,
        "description": recipe.description,
        "version": recipe.version,
        "status": recipe.status,
        "created_by": recipe.created_by,
        "created_at": recipe.created_at,
        "no_of_materials": recipe.no_of_materials
    }
    end_time = time.time()
    execution_time_ms = round((end_time - start_time) * 1000, 2)

    return jsonify({**result, "execution_time_ms": execution_time_ms})

@recipe_bp.route("/recipes/<int:recipe_id>", methods=["PUT"])
def update_recipe(recipe_id):
    start_time = time.time()
    try:
        recipe = Recipe.query.get(recipe_id)
        if not recipe:
            return jsonify({"message": "Recipe not found"}), 404

        data = request.get_json()
        recipe.name = data.get("name", recipe.name)
        recipe.code = data.get("code", recipe.code)
        recipe.description = data.get("description", recipe.description)
        recipe.version = data.get("version", recipe.version)
        recipe.status = data.get("status", recipe.status)
        recipe.no_of_materials = data.get("no_of_materials", recipe.no_of_materials)
        if "sequence" in data:
            recipe.sequence = data["sequence"]

        db.session.commit()
        end_time = time.time()
        return jsonify({"message": "Recipe updated successfully", "execution_time_ms": round((end_time - start_time) * 1000, 2)}), 200

    except Exception as e:
        db.session.rollback()
        logging.error(f"Error updating recipe {recipe_id}: {str(e)}")
        return jsonify({"message": "An error occurred while updating the recipe."}), 500

@recipe_bp.route("/recipes/<int:recipe_id>", methods=["DELETE"])
def delete_recipe(recipe_id):
    start_time = time.time()
    try:
        db.session.query(ProductionOrder).filter(ProductionOrder.recipe_id == recipe_id).delete(synchronize_session=False)
        db.session.query(RecipeMaterial).filter(RecipeMaterial.recipe_id == recipe_id).delete(synchronize_session=False)
        recipe = Recipe.query.get(recipe_id)
        if not recipe:
            return jsonify({"message": "Recipe not found"}), 404

        db.session.delete(recipe)
        db.session.commit()
        end_time = time.time()
        return jsonify({"message": "Recipe and related data deleted successfully", "execution_time_ms": round((end_time - start_time) * 1000, 2)}), 200

    except IntegrityError as e:
        db.session.rollback()
        return jsonify({"message": "Integrity error, check related records for consistency"}), 400
    except Exception as e:
        db.session.rollback()
        return jsonify({"message": "An error occurred while deleting the recipe"}), 500

@recipe_bp.route("/recipe_materials", methods=["POST"])
def create_or_update_recipe_material():
    start_time = time.time()
    try:
        data = request.get_json()
        if not data:
            raise BadRequest("No input data provided.")

        recipe_id = data.get("recipe_id")
        material_id = data.get("material_id")
        set_point = data.get("set_point")
        status = data.get("status")
        bucket_id = data.get("bucket_id")
        use_scale = data.get("use_scale", False)

        if use_scale:
            from models.scale import ScaleClient
            scale_client = ScaleClient()
            actual = scale_client.get_net_weight()
            if actual is None:
                return jsonify({"error": "Failed to read weight from scale."}), 500
        else:
            actual = data.get("actual")

        # ✅ Input validations
        if not recipe_id or not material_id or set_point is None or not status:
            raise BadRequest("Missing required fields.")
        if actual is None:
            raise BadRequest("Actual weight is required.")
        if not isinstance(recipe_id, int) or not isinstance(material_id, int):
            raise BadRequest("recipe_id and material_id must be integers.")
        if not isinstance(set_point, (int, float)) or not isinstance(actual, (int, float)):
            raise BadRequest("set_point and actual must be numeric.")

        # ✅ Validate bucket
        if bucket_id is not None:
            from models import StorageBucket
            if not StorageBucket.query.get(bucket_id):
                raise BadRequest("Invalid bucket_id.")

        # ✅ Use provided margin if exists, else calculate
        if "margin" in data:
            margin = float(data["margin"])
        else:
            margin = 0.0 if set_point == 0 else round(((float(set_point) - float(actual)) / float(set_point)) * 100, 2)

        # ✅ Check if material already exists
        existing = RecipeMaterial.query.filter_by(recipe_id=recipe_id, material_id=material_id).first()

        if existing:
            existing.set_point = set_point
            existing.actual = actual
            existing.margin = margin
            existing.status = status
            existing.bucket_id = bucket_id
            db.session.commit()

            socketio.emit("recipe_material_updated", {
                "recipe_id": recipe_id,
                "material_id": material_id,
                "actual": actual,
                "set_point": set_point,
                "status": status,
                "bucket_id": bucket_id,
                "margin": margin,
                "type": "update"
            })

            return jsonify({
                "message": "Recipe material updated successfully!",
                "execution_time_ms": round((time.time() - start_time) * 1000, 2)
            }), 200

        # ✅ New entry
        new_material = RecipeMaterial(
            recipe_id=recipe_id,
            material_id=material_id,
            set_point=set_point,
            actual=actual,
            margin=margin,
            status=status,
            bucket_id=bucket_id
        )
        db.session.add(new_material)
        db.session.commit()

        socketio.emit("recipe_material_created", {
            "recipe_id": recipe_id,
            "material_id": material_id,
            "actual": actual,
            "set_point": set_point,
            "status": status,
            "bucket_id": bucket_id,
            "margin": margin,
            "type": "create"
        })

        return jsonify({
            "message": "Recipe material created successfully!",
            "execution_time_ms": round((time.time() - start_time) * 1000, 2)
        }), 201

    except BadRequest as e:
        logging.error(f"Bad request: {e.description}")
        return jsonify({"error": e.description}), 400

    except Exception as e:
        logging.error(f"Unexpected error: {str(e)}")
        return jsonify({"error": "An unexpected error occurred."}), 500

@recipe_bp.route("/recipe_materials/active", methods=["GET"])
def get_active_recipe_material():
    try:
        # ✅ Step 1: Get latest verified production order with pending materials
        latest_order = (
            db.session.query(ProductionOrder)
            .join(RecipeMaterial, ProductionOrder.recipe_id == RecipeMaterial.recipe_id)
            .filter(
                RecipeMaterial.status == "pending",
                ProductionOrder.status == "verified"
            )
            .order_by(ProductionOrder.created_at.desc())
            .first()
        )

        if not latest_order:
            return jsonify({"message": "No verified orders with pending recipe materials."}), 200

        # ✅ Step 2: Fetch all pending materials for this recipe
        pending_materials = (
            db.session.query(RecipeMaterial)
            .options(joinedload(RecipeMaterial.material), joinedload(RecipeMaterial.recipe))
            .filter(
                RecipeMaterial.recipe_id == latest_order.recipe_id,
                RecipeMaterial.status == "pending"
            )
            .order_by(RecipeMaterial.recipe_material_id)
            .all()
        )

        if not pending_materials:
            return jsonify({"message": "No pending materials for this verified recipe."}), 200

        # ✅ Step 3: Serialize response data
        materials_data = []
        for mat in pending_materials:
            material = mat.material
            recipe = mat.recipe
            if not material or not recipe:
                continue

            materials_data.append({
                "recipe_id": recipe.recipe_id,
                "recipe_name": recipe.name,
                "material_id": mat.material_id,
                "material_name": material.title,
                "barcode": material.barcode_id,
                "set_point": mat.set_point,
                "actual": mat.actual,
                "margin": mat.margin,
                "status": mat.status,
                "bucket_id": mat.bucket_id
            })

        # ✅ Step 4: Emit all active materials via WebSocket
        socketio.emit("active_recipe_materials", {
            "recipe_id": latest_order.recipe_id,
            "recipe_name": latest_order.recipe.name,
            "materials": materials_data
        }, namespace='/')

        # ✅ Step 5: Return first pending material (backward compatibility)
        return jsonify(materials_data[0]), 200

    except Exception as e:
        logging.error(f"❌ Error in get_active_recipe_material: {str(e)}", exc_info=True)
        return jsonify({"error": "Internal server error"}), 500


import requests
@recipe_bp.route("/recipe_materials/weigh-and-update", methods=["POST"])
def weigh_and_update_material():
    try:
        from extensions import socketio

        # ✅ Step 1: Read live scale weight
        try:
            # Uncomment for real scale:
            scale_response = requests.get("http://127.0.0.1:5000/api/scale/read-weight", timeout=5)
            scale_response.raise_for_status()
            scale_data = scale_response.json()
            current_weight = scale_data["data"].get("weight_kg")
            # current_weight = 20  # 🧪 Mocked for development
        except Exception as e:
            logging.error(f"❌ Scale read failed: {str(e)}", exc_info=True)
            return jsonify({"success": False, "message": "Unable to read from scale"}), 500

        # ✅ Step 2: Get the active verified production order
        verified_order = (
            db.session.query(ProductionOrder)
            .filter(ProductionOrder.status == "verified")
            .order_by(ProductionOrder.created_at.desc())  # optional if multiple verified not expected
            .first()
        )

        if not verified_order:
            return jsonify({"success": False, "message": "No verified production order found"}), 200

        recipe_id = verified_order.recipe_id

        # ✅ Step 3: Fetch all materials for the active recipe
        materials = (
            db.session.query(RecipeMaterial)
            .filter(RecipeMaterial.recipe_id == recipe_id)
            .order_by(RecipeMaterial.recipe_material_id)
            .all()
        )

        # ✅ Step 4: Loop to find first pending material
        for index, material in enumerate(materials):
            if material.status != "pending":
                continue

            # ✅ Ensure previous material is dosed
            if index > 0 and materials[index - 1].status != "Dosed":
                return jsonify({
                    "success": False,
                    "message": "Waiting for previous material to be dosed"
                }), 200

            # ✅ Weight margin logic
            set_point = float(material.set_point or 0)
            margin_g = float(material.margin or 0)
            margin_kg = margin_g / 1000
            lower_limit = set_point - margin_kg

            if current_weight >= lower_limit:
                # ✅ Check if this is the final pending material
                remaining_before = db.session.query(RecipeMaterial).filter_by(
                    recipe_id=material.recipe_id, status="pending"
                ).count()
                is_final_material = (remaining_before == 1)

                # ✅ Mark material as dosed
                material.actual = current_weight
                material.margin = round((current_weight - set_point) * 1000, 2)
                material.status = "Dosed"
                db.session.commit()

                # ✅ Log to dosed table
                batch_size = float(verified_order.batch_size) if verified_order and verified_order.batch_size else 1

                dosed_entry = DosedRecipeMaterial(
                    recipe_id=material.recipe_id,
                    material_id=material.material_id,
                    set_point=set_point,
                    actual=current_weight,
                    margin=material.margin,
                    batch_size=batch_size
                )
                db.session.add(dosed_entry)
                db.session.commit()

                reset_done = False

                # ✅ If this was the final material, reset the recipe
                if is_final_material:
                    all_materials = db.session.query(RecipeMaterial).filter_by(recipe_id=recipe_id).all()
                    for m in all_materials:
                        if m.set_point and batch_size:
                            m.set_point = m.set_point / batch_size
                        m.status = "pending"
                        m.actual = None
                        m.margin = None
                    db.session.commit()
                    reset_done = True

                    socketio.emit("recipe_reset", {
                        "recipe_id": recipe_id,
                        "message": "All materials dosed and reset for next round."
                    })

                # ✅ Emit material update
                socketio.emit("material_updated", {
                    "recipe_id": recipe_id,
                    "material_id": material.material_id,
                    "material_name": material.material.title,
                    "status": "Dosed",
                    "actual": current_weight,
                    "set_point": set_point,
                    "deviation": material.margin
                })

                return jsonify({
                    "success": True,
                    "message": "Dosed successfully",
                    "reset_done": reset_done,
                    "total_remaining": 0 if reset_done else (remaining_before - 1),
                    "data": {
                        "recipe_material_id": material.recipe_material_id,
                        "material_id": material.material_id,
                        "material_name": material.material.title,
                        "set_point": set_point,
                        "actual": current_weight,
                        "margin": material.margin,
                        "status": "Dosed"
                    }
                }), 200

            else:
                # ❌ Weight too low
                return jsonify({
                    "success": False,
                    "reason": "underweight",
                    "message": "Weight not yet sufficient",
                    "data": {
                        "material_id": material.material_id,
                        "material_name": material.material.title,
                        "actual": current_weight,
                        "set_point": set_point,
                        "margin_g": margin_g
                    }
                }), 200

        return jsonify({"success": False, "message": "No pending materials found"}), 200

    except Exception as e:
        logging.error(f"❌ Fatal error in weigh-and-update: {str(e)}", exc_info=True)
        return jsonify({"success": False, "message": "Internal server error"}), 500

@recipe_bp.route("/recipe_materials", methods=["GET"])
def get_recipe_materials():
    start_time = time.time()
    materials = RecipeMaterial.query.all()
    result = [
        {
            "recipe_material_id": mat.recipe_material_id,
            "recipe_id": mat.recipe_id,
            "material_id": mat.material_id,
            "set_point": str(mat.set_point) if mat.set_point is not None else None,
            "actual": str(mat.actual) if mat.actual is not None else None,
            "margin": str(mat.margin) if mat.margin is not None else None
        }
        for mat in materials
    ]
    end_time = time.time()
    return jsonify({"materials": result, "execution_time_ms": round((end_time - start_time) * 1000, 2)})

@recipe_bp.route("/recipe_materials/<int:recipe_material_id>", methods=["PUT"])
def update_recipe_material(recipe_material_id):
    start_time = time.time()
    material = RecipeMaterial.query.get(recipe_material_id)
    if not material:
        return jsonify({"message": "Recipe material not found"}), 404

    data = request.get_json()
    material.material_id = data.get("material_id", material.material_id)
    material.set_point = data.get("set_point", material.set_point)

    db.session.commit()
    end_time = time.time()
    return jsonify({"message": "Recipe material updated successfully", "execution_time_ms": round((end_time - start_time) * 1000, 2)})

@recipe_bp.route("/recipe_materials/<int:recipe_id>", methods=["GET"])
def get_recipe_materials_by_recipe_id(recipe_id):
    start_time = time.time()
    materials = RecipeMaterial.query.filter_by(recipe_id=recipe_id).all()
    if not materials:
        return jsonify({"message": "No materials found for this recipe"}), 404

    result = [
        {
            "recipe_material_id": mat.recipe_material_id,
            "recipe_id": mat.recipe_id,
            "material_id": mat.material_id,
            "set_point": str(mat.set_point) if mat.set_point is not None else None,
            "actual": str(mat.actual) if mat.actual is not None else None,
            "margin": str(mat.margin) if mat.margin is not None else None
        }
        for mat in materials
    ]
    end_time = time.time()
    return jsonify({"materials": result, "execution_time_ms": round((end_time - start_time) * 1000, 2)})

@recipe_bp.route("/recipe_materials/<int:recipe_material_id>", methods=["DELETE"])
def delete_recipe_material(recipe_material_id):
    start_time = time.time()
    material = RecipeMaterial.query.get(recipe_material_id)
    if not material:
        return jsonify({"message": "Recipe material not found"}), 404

    db.session.delete(material)
    db.session.commit()
    end_time = time.time()
    return jsonify({"message": "Recipe material deleted successfully", "execution_time_ms": round((end_time - start_time) * 1000, 2)})
@recipe_bp.route("/recipe_materials/dosed", methods=["GET"])
def get_dosed_materials():
    try:
        page = int(request.args.get("page", 1))
        per_page = int(request.args.get("per_page", 20))
        barcode = request.args.get("barcode")

        query = RecipeMaterial.query.filter(RecipeMaterial.status.in_(["Dosed", "Rejected"])) \
            .join(RecipeMaterial.recipe) \
            .join(RecipeMaterial.material)

        if barcode:
            from models.material import Material  # adjust if needed
            query = query.filter(Material.barcode_id == barcode)

        query = query.order_by(RecipeMaterial.recipe_material_id)
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)

        result = [
            {
                "recipe_material_id": m.recipe_material_id,
                "recipe_name": m.recipe.name if m.recipe else None,
                "material_name": m.material.title if m.material else None,
                "set_point": str(m.set_point) if m.set_point is not None else None,
                "actual": str(m.actual) if m.actual is not None else None,
                "margin": str(m.margin) if m.margin is not None else None,
                "status": m.status
            }
            for m in pagination.items
        ]

        return jsonify({
            "materials": result,
            "page": pagination.page,
            "pages": pagination.pages,
            "per_page": pagination.per_page,
            "total": pagination.total
        }), 200

    except Exception as e:
        import traceback
        print("❌ ERROR in /recipe_materials/dosed")
        print(traceback.format_exc())
        return jsonify({"error": "Internal Server Error", "details": str(e)}), 500
@recipe_bp.route("/recipe_materials/bypass/<int:recipe_id>", methods=["POST"])
def bypass_pending_materials(recipe_id):
    try:
        # Step 1: Get all 'pending' recipe materials for the given recipe_id
        pending_materials = (
            db.session.query(RecipeMaterial)
            .filter(RecipeMaterial.recipe_id == recipe_id, RecipeMaterial.status == 'pending')
            .all()
        )

        if not pending_materials:
            return jsonify({"message": f"No pending materials found for recipe ID {recipe_id}."}), 200

        bypassed_ids = []
        for mat in pending_materials:
            mat.status = 'rejected'
            bypassed_ids.append(mat.recipe_material_id)

            # Optionally log or emit WebSocket for each bypassed material
            socketio.emit('recipe_material_updated', {
                "material_id": mat.material_id,
                "status": "Rejected"
            }, namespace='/')

        db.session.commit()

        logging.info(f"Bypassed {len(bypassed_ids)} materials for recipe ID {recipe_id}.")
        return jsonify({
            "message": f"{len(bypassed_ids)} materials bypassed.",
            "bypassed_ids": bypassed_ids
        }), 200

    except Exception as e:
        db.session.rollback()
        logging.error(f"Error bypassing materials for recipe ID {recipe_id}: {str(e)}")
        return jsonify({"error": "Internal server error."}), 500

@recipe_bp.route("/dosed_recipe_materials", methods=["GET"])
def get_dosed_recipe_materials():
    try:
        # Get pagination parameters from query string
        page = int(request.args.get("page", 1))
        per_page = int(request.args.get("per_page", 10))

        query = DosedRecipeMaterial.query.order_by(DosedRecipeMaterial.dosed_at.desc())

        # Apply pagination
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)

        records = []
        for record in pagination.items:
            records.append({
                "id": record.id,
                "recipe_id": record.recipe_id,
                "recipe_name": record.recipe.name if record.recipe else "—",
                "material_id": record.material_id,
                "material_name": record.material.title if record.material else "—",
                "set_point": record.set_point,
                "actual": record.actual,
                "margin": record.margin,
                "batch_size": record.batch_size,
                "dosed_at": record.dosed_at.strftime("%Y-%m-%d %H:%M:%S") if record.dosed_at else None
            })

        return jsonify({
            "page": pagination.page,
            "per_page": pagination.per_page,
            "total": pagination.total,
            "pages": pagination.pages,
            "records": records
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Internal server error: {str(e)}"}), 500

@recipe_bp.route("/dosed_recipe_materials/delete-all", methods=["DELETE"])
# @jwt_required(locations=["headers"])
# @role_required(["admin"])  # 🔐 only allow admins
def delete_all_dosed_recipe_materials():
    try:
        num_deleted = db.session.query(DosedRecipeMaterial).delete()
        db.session.commit()

        return jsonify({
            "message": f"✅ Successfully deleted {num_deleted} dosed recipe material records."
        }), 200

    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({
            "error": f"❌ Failed to delete records: {str(e)}"
        }), 500
